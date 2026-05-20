import re
from pathlib import Path

import openpyxl
import pandas as pd

from .config import COL_SUFFIX_RE, TICKER_RE

CANONICAL_DATA_COLS = [
    'Date',
    'Last Price',
    'Bid Price',
    'Ask Price',
    'Mid Price',
    'Implied Volatility Mid',
    'Volume',
    'Open Interest',
    'Underlying Price',
    'Delta Mid Price',
    'Gamma Mid',
]

META_COLS = ['ticker', 'option_type', 'strike', 'expiry', 'sheet_name']
PANEL_COLS = CANONICAL_DATA_COLS + META_COLS

def parse_ticker(ticker: str) -> dict:
    m = re.search(TICKER_RE, ticker)
    if not m:
        raise ValueError(f"Cannot parse ticker: {ticker!r}")
    expiry_str, opt_type, strike = m.group(1), m.group(2), int(m.group(3))
    expiry = pd.to_datetime(expiry_str, format='%m/%d/%y')
    return {'expiry': expiry, 'option_type': opt_type, 'strike': strike}

def normalize_col(col) -> str:
    if not isinstance(col, str):
        return str(col)
    return re.sub(COL_SUFFIX_RE, '', col).strip()

def find_header_row(ws, max_search: int = 60) -> int | None:
    for i in range(1, max_search + 1):
        for j in range(1, 6):
            v = ws.cell(row=i, column=j).value
            if isinstance(v, str) and v.strip().lower() == 'date':
                return i
    return None

def find_ticker(ws, max_search: int = 5) -> str:
    for i in range(1, max_search + 1):
        for j in range(1, 6):
            v = ws.cell(row=i, column=j).value
            if isinstance(v, str) and 'SPX US' in v and 'Index' in v:
                return v.strip()
    raise ValueError("No 'SPX US ... Index' ticker found in first 5 rows")

def load_sheet(ws, sheet_name: str) -> pd.DataFrame:
    ticker = find_ticker(ws)
    meta = parse_ticker(ticker)

    hdr_row = find_header_row(ws)
    if hdr_row is None:
        raise ValueError(f"No 'Date' header row in sheet {sheet_name!r}")

    max_col = ws.max_column
    headers = []
    seen = {}
    for j in range(1, max_col + 1):
        v = ws.cell(row=hdr_row, column=j).value
        h = normalize_col(v) if v is not None else f'_col{j}'

        if h in seen:
            seen[h] += 1
            h = f"{h}_{seen[h]}"
        else:
            seen[h] = 0
        headers.append(h)

    data = []
    for i in range(hdr_row + 1, ws.max_row + 1):
        row_vals = [ws.cell(row=i, column=j).value for j in range(1, max_col + 1)]

        if all(v is None for v in row_vals):
            continue

        if row_vals[0] is None:
            continue
        data.append(row_vals)

    df = pd.DataFrame(data, columns=headers)

    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    df = df.dropna(subset=['Date']).reset_index(drop=True)

    df = df.sort_values('Date').reset_index(drop=True)

    df['ticker'] = ticker
    df['option_type'] = meta['option_type']
    df['strike'] = meta['strike']
    df['expiry'] = meta['expiry']
    df['sheet_name'] = sheet_name

    return df

def load_workbook(xlsx_path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=False, data_only=True)
    dfs = []
    skipped = []
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() == 'spx':
            skipped.append(sheet_name)
            continue
        ws = wb[sheet_name]
        try:
            df = load_sheet(ws, sheet_name)
            dfs.append(df)
        except Exception as e:
            print(f"  ! Sheet {sheet_name!r} failed: {e}")
            raise
    wb.close()
    if not dfs:
        raise RuntimeError("No option sheets parsed")
    panel = pd.concat(dfs, ignore_index=True, sort=False)

    dup = panel.duplicated(subset=['Date', 'ticker'])
    if dup.any():
        n_dup = dup.sum()
        sample = panel[dup].head(5)[['Date', 'ticker']]
        raise RuntimeError(f"Duplicate (Date, ticker) rows: {n_dup}\n{sample}")

    keep = [c for c in PANEL_COLS if c in panel.columns]
    panel = panel[keep].copy()

    ordered = META_COLS + [c for c in CANONICAL_DATA_COLS if c in panel.columns]
    panel = panel[ordered].sort_values(['ticker', 'Date']).reset_index(drop=True)

    return panel
